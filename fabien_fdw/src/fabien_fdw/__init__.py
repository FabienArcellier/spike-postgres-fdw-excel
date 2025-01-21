import csv
import io
import logging
import openpyxl
from logging import ERROR

from multicorn import ForeignDataWrapper
from multicorn.utils import log_to_postgres


class ExcelForeignDataWrapper(ForeignDataWrapper):
    """
    https://github.com/pgsql-io/multicorn2
    documentation: https://multicorn.org/
    """

    def __init__(self, options, columns):
        super(ExcelForeignDataWrapper, self).__init__(options, columns)
        self.options = options
        self.columns = columns
        self._row_id_column = options.get('rowid_column', None)


    def execute(self, quals, columns):
        """

        :param quals:
        :param columns:
        :return:
        """
        try:
            workbook = openpyxl.load_workbook(self.options.get('path'), data_only=True)
            sheet = workbook[self.options.get('sheet')]

            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            column_indices = [headers.index(col) for col in columns]

            for row in sheet.iter_rows(min_row=2):
                row_data = {}
                for col, idx in zip(columns, column_indices):
                    row_data[col] = row[idx].value
                    if row_data[col] == "Oui":
                        row_data[col] = True
                    elif row_data[col] == "Non":
                        row_data[col] = False

                yield row_data

        except Exception as e:
            raise Exception(f"Error reading Excel file: {e}") from e

    def insert(self, new_values):
        log_to_postgres("Executing insert operation", level=ERROR)
        try:
            workbook = openpyxl.load_workbook(self.options.get('path'))
            sheet = workbook[self.options.get('sheet')]
            row_data = [new_values.get(col, "") for col in self.columns.keys()]
            sheet.append(row_data)
            workbook.save(self.options.get('path'))
        except Exception as e:
            raise Exception(f"Error inserting into Excel file: {str(e)}") from e

    def update(self, row_id, new_values):
        """
        Updates an existing row in the Excel sheet.
        """
        raise NotImplementedError("Update operation is not supported")

    def delete(self, row_id):
        raise NotImplementedError("Delete operation is not supported")

    @property
    def rowid_column(self):
        return self._row_id_column